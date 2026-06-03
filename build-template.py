from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

RED="9E1B32"; BLUE="0A3D62"; GOLD="F5B301"; LITE="FDEEF0"; HEADTXT="FFFFFF"
HFONT=Font(name="Arial", bold=True, color=HEADTXT, size=11)
TFONT=Font(name="Arial", size=10)
NOTE=Font(name="Arial", size=9, italic=True, color="7A6A64")
thin=Side(style="thin", color="D9D9D9")
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=Workbook()

def style_headers(ws, ncol, fill=RED, freeze="A3"):
    for c in range(1, ncol+1):
        cell=ws.cell(row=2, column=c)
        cell.font=HFONT; cell.fill=PatternFill("solid", fgColor=fill)
        cell.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border=BORDER
    ws.row_dimensions[2].height=34
    ws.freeze_panes=freeze

def put_title(ws, text, span):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=span)
    t=ws.cell(row=1,column=1,value=text)
    t.font=Font(name="Arial", bold=True, size=13, color=RED)
    t.alignment=Alignment(vertical="center")
    ws.row_dimensions[1].height=24

def write_rows(ws, rows, start=3):
    for i,row in enumerate(rows):
        for j,val in enumerate(row):
            c=ws.cell(row=start+i, column=j+1, value=val)
            c.font=TFONT; c.border=BORDER
            c.alignment=Alignment(vertical="top", wrap_text=True)

# ============ 1) HƯỚNG DẪN ============
ws=wb.active; ws.title="HuongDan"
ws.column_dimensions["A"].width=4
ws.column_dimensions["B"].width=110
ws.sheet_view.showGridLines=False
lines=[
 ("BK360 — TEMPLATE NHẬP LIỆU 2D MAP", "title"),
 ("Điền dữ liệu vào các sheet bên dưới rồi gửi lại file này KÈM thư mục ảnh/audio. Hệ thống sẽ tự load để dựng map.",""),
 ("",""),
 ("CÁC SHEET:", "h"),
 ("• DiaDiem  — thông tin từng địa điểm (tên, mô tả, file ảnh xưa/nay, file 360°, file audio…)",""),
 ("• LichSu   — dòng thời gian lịch sử của từng địa điểm (mỗi mốc 1 dòng)",""),
 ("• SuKien   — các sự kiện/chiến dịch (vd: Sự kiện 70 năm, Khai giảng, Ngày trở về)",""),
 ("• LichTrinh— lịch trình hoạt động của mỗi sự kiện (giờ, địa điểm, đang diễn ra…)",""),
 ("",""),
 ("QUY TẮC QUAN TRỌNG:", "h"),
 ("1) Cột 'id' là MÃ DUY NHẤT, viết thường không dấu (vd: c1, library, b1). Dùng để liên kết giữa các sheet và đặt tên file.",""),
 ("2) Cột link_audio / link_anh_xua / link_anh_nay / link_anh_360: DÁN NGUYÊN LINK (URL) tới ảnh/audio đã tải lên. Bỏ trống nếu chưa có.",""),
 ("3) Dán LINK CHIA SẺ của Google Drive HOẶC OneDrive cũng được — hệ thống TỰ chuyển sang link nhúng trực tiếp.",""),
 ("   BẮT BUỘC đặt quyền 'Bất kỳ ai có liên kết đều xem được' (Anyone with the link). Link riêng tư sẽ không hiển thị.",""),
 ("4) Ảnh 360° là ảnh equirectangular (tỉ lệ 2:1). Ảnh xưa/nay nên cùng khung hình để slider so sánh khớp.",""),
 ("5) Cột 'loai': chọn 'spot' (địa điểm thường) hoặc 'event' (điểm sự kiện).",""),
 ("6) Cột 'hien': 'Có' = hiện trên map, 'Không' = ẩn.",""),
 ("7) map_x / map_y: VỊ TRÍ HOTSPOT THEO % ẢNH NỀN (0–100). map_x: mép trái=0, phải=100. map_y: trên=0, dưới=100.",""),
 ("   → Mở app, bật nút 'Lấy toạ độ', bấm lên ảnh nền để lấy nhanh giá trị %.",""),
 ("8) 'thuyet_minh_vi': lời thuyết minh dạng chữ (đọc bằng máy). Nếu có file audio thật thì điền 'file_audio' và để trống cột này cũng được.",""),
 ("",""),
 ("BÀN GIAO:", "h"),
 ("→ Gửi lại: CHỈ cần file Excel này (ảnh/audio đã nằm trên link công khai). Không cần đính kèm file.",""),
 ("→ Mình sẽ load file → sinh dữ liệu → map tự hiển thị ảnh/giọng nói/lịch sử/sự kiện.",""),
]
r=1
for text,kind in lines:
    cell=ws.cell(row=r, column=2, value=text)
    if kind=="title":
        cell.font=Font(name="Arial", bold=True, size=15, color=RED)
    elif kind=="h":
        cell.font=Font(name="Arial", bold=True, size=11, color=BLUE)
    else:
        cell.font=Font(name="Arial", size=10.5, color="333333")
    cell.alignment=Alignment(wrap_text=True, vertical="center")
    r+=1

# ============ 2) DIA DIEM ============
ws=wb.create_sheet("DiaDiem")
ws.sheet_view.showGridLines=False
cols=[("id",12),("ten_vi",26),("ten_en",24),("nhan_ngan",16),("nam",20),("loai",10),
      ("map_x",8),("map_y",8),("mo_ta_vi",46),("mo_ta_en",40),("thuyet_minh_vi",46),
      ("link_audio",30),("link_anh_xua",30),("link_anh_nay",30),("link_anh_360",30),("hien",8)]
put_title(ws,"DIA DIEM — thông tin hiển thị từng toà nhà / điểm",len(cols))
for j,(name,wd) in enumerate(cols):
    ws.cell(row=2,column=j+1,value=name); ws.column_dimensions[get_column_letter(j+1)].width=wd
style_headers(ws,len(cols))

# map_x/map_y = % theo ảnh nền (0–100)
D=[
 ["parabol","Cổng Parabol","Parabol Gate","Biểu tượng","Khánh thành 1965","event",8,57,
  "Cổng Parabol mái vòm 'ngọn sóng' — biểu tượng của Bách Khoa Hà Nội (rộng 8m, cao 12,5m).","",
  "Cổng Parabol là biểu tượng kiến trúc gắn liền với Bách Khoa Hà Nội.","","","","","Có"],
 ["c1","Nhà C1 (toà trung tâm)","Building C1","Trung tâm","Số 1 Đại Cồ Việt","spot",34,28,
  "Toà nhà trung tâm với tháp lam chắn nắng, mái 'ngọn sóng' trên đỉnh và logo BK; phía trước là Quảng trường C1 (I ❤ BK).","",
  "Toà nhà C1 — trung tâm hành chính của trường.",
  "https://drive.google.com/file/d/AUDIO_FILE_ID/view?usp=sharing",
  "https://drive.google.com/file/d/XUA_FILE_ID/view?usp=sharing",
  "https://drive.google.com/file/d/NAY_FILE_ID/view?usp=sharing",
  "https://drive.google.com/file/d/PANO_FILE_ID/view?usp=sharing","Có"],
 ["c2","Hội trường C2","C2 Hall","Hội trường","Trường Vật liệu / Điện-ĐT / CNTT&TT","spot",20,42,
  "Hội trường C2 — nơi tổ chức sự kiện, hội thảo; khu các trường Vật liệu, Điện-Điện tử, CNTT&TT.","","","","","","","Có"],
 ["c7","Nhà C7","Building C7","Khối chữ E","Viện KT&QL; Vật lý KT; KH&CN GD; Ngoại ngữ","spot",62,50,
  "Toà C7 hình chữ E — Viện Kinh tế & Quản lý (109), Vật lý Kỹ thuật (111), KH&CN Giáo dục (218), Ngoại ngữ (M317).","","","","","","","Có"],
 ["library","Thư viện Tạ Quang Bửu","Ta Quang Buu Library","Thư viện","Trung tâm tri thức","spot",52,67,
  "Thư viện Tạ Quang Bửu — khối nhà nhiều tầng có đồng hồ mặt tiền, trung tâm tri thức của trường.","",
  "Thư viện Tạ Quang Bửu — trái tim tri thức của Bách Khoa.","",
  "https://1drv.ms/i/s!AbcXuaExample","https://1drv.ms/i/s!AbcNayExample","","Có"],
 ["b1","Nhà B1","Building B1","Trường Hoá","Trường Hoá & Khoa học Sự sống","spot",78,67,
  "Toà B1 — tháp cao mặt kính, sảnh kính tròn phía trước; Trường Hoá và Khoa học Sự sống.","","","","","","","Có"],
 ["c9","Nhà C9","Building C9","Toà C9","","spot",24,54,"","","","","","","","Có"],
 ["c10","Nhà C10 (ITIMS)","Building C10","ITIMS","","spot",51,54,"","","","","","","","Có"],
 ["b5","KTX B5","Dorm B5","Ký túc xá","","spot",88,37,"","","","","","","","Có"],
 ["b9","KTX B9","Dorm B9","Ký túc xá","","spot",88,45,"","","","","","","","Có"],
 ["d2a","Nhà D2A","Building D2A","Toà D2A","","spot",18,64,"","","","","","","","Có"],
 ["vietduc","Trung tâm Việt-Đức","Vietnam-Germany Center","TT Việt-Đức","Khoa Toán-Tin","spot",34,64,
  "Trung tâm Việt-Đức / Khoa Toán-Tin.","","","","","","","Có"],
]
write_rows(ws,D)
# validation
dv_loai=DataValidation(type="list", formula1='"spot,event"', allow_blank=True); ws.add_data_validation(dv_loai); dv_loai.add(f"F3:F500")
dv_hien=DataValidation(type="list", formula1='"Có,Không"', allow_blank=True); ws.add_data_validation(dv_hien); dv_hien.add(f"P3:P500")

# ============ 3) LICH SU ============
ws=wb.create_sheet("LichSu")
ws.sheet_view.showGridLines=False
cols=[("dia_diem_id",16),("nam",16),("noi_dung",90),("thu_tu",8)]
put_title(ws,"LICH SU — dòng thời gian từng địa điểm (mỗi mốc 1 dòng)",len(cols))
for j,(name,wd) in enumerate(cols):
    ws.cell(row=2,column=j+1,value=name); ws.column_dimensions[get_column_letter(j+1)].width=wd
style_headers(ws,len(cols),fill=BLUE)
L=[
 ["c1","1956","Trường ĐH Bách khoa Hà Nội thành lập (06/3/1956) — trường kỹ thuật đầu tiên của Việt Nam.",1],
 ["c1","2022","Chuyển thành Đại học Bách khoa Hà Nội (mô hình đại học có trường thành viên).",2],
 ["library","1956","Tủ sách kỹ thuật phục vụ giảng dạy từ những ngày đầu.",1],
 ["library","2006","Tòa thư viện hiện đại khánh thành, mang tên GS. Tạ Quang Bửu.",2],
 ["library","2020","Số hóa kho học liệu, phát triển không gian học tập sáng tạo.",3],
 ["parabol","1960s","Cổng Parabol hình thành trên đường Đại Cồ Việt, trở thành biểu tượng của trường.",1],
]
write_rows(ws,L)

# ============ 4) SU KIEN ============
ws=wb.create_sheet("SuKien")
ws.sheet_view.showGridLines=False
cols=[("id",16),("ten_vi",26),("ten_en",24),("icon",8),("bat",8)]
put_title(ws,"SU KIEN — các chiến dịch sự kiện (bật/tắt)",len(cols))
for j,(name,wd) in enumerate(cols):
    ws.cell(row=2,column=j+1,value=name); ws.column_dimensions[get_column_letter(j+1)].width=wd
style_headers(ws,len(cols),fill=GOLD)
for c in range(1,len(cols)+1):
    ws.cell(row=2,column=c).font=Font(name="Arial",bold=True,color="3A2B00",size=11)
S=[
 ["anniv70","Sự kiện 70 năm","70th Anniversary","⭐","Có"],
 ["homecoming","BK – Ngày trở về","BK Homecoming","🎓","Không"],
 ["khaigiang","Lễ Khai giảng","Opening Ceremony","🎉","Không"],
]
write_rows(ws,S)
dv_bat=DataValidation(type="list", formula1='"Có,Không"', allow_blank=True); ws.add_data_validation(dv_bat); dv_bat.add("E3:E200")

# ============ 5) LICH TRINH ============
ws=wb.create_sheet("LichTrinh")
ws.sheet_view.showGridLines=False
cols=[("su_kien_id",16),("gio",10),("dia_diem_id",16),("dang_dien_ra",14),("tieu_de_vi",40),("tieu_de_en",36)]
put_title(ws,"LICH TRINH — hoạt động của mỗi sự kiện",len(cols))
for j,(name,wd) in enumerate(cols):
    ws.cell(row=2,column=j+1,value=name); ws.column_dimensions[get_column_letter(j+1)].width=wd
style_headers(ws,len(cols),fill=BLUE)
T=[
 ["anniv70","08:00","stadium","Có","Lễ khai mạc 70 năm","70th opening ceremony"],
 ["anniv70","09:30","library","Có",'Triển lãm "Xưa & Nay"',"Then & Now exhibition"],
 ["anniv70","10:30","c2","Không","Tọa đàm cựu sinh viên","Alumni talk"],
 ["anniv70","14:00","c1","Không","Tham quan phòng truyền thống","Tradition room tour"],
 ["anniv70","19:30","stadium","Không","Đêm Gala hội ngộ","Gala night"],
 ["homecoming","08:30","parabol","Không","Đón tiếp cựu sinh viên","Alumni welcome"],
]
write_rows(ws,T)
dv_live=DataValidation(type="list", formula1='"Có,Không"', allow_blank=True); ws.add_data_validation(dv_live); dv_live.add("D3:D500")

wb.save("BK360-Template-2Dmap.xlsx")
print("WROTE BK360-Template-2Dmap.xlsx")
